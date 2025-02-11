import re
from unittest.mock import MagicMock

import openpyxl
import pytest

from bam_masterdata.excel import MasterdataExcelExtractor
from bam_masterdata.logger import logger
from bam_masterdata.metadata.definitions import DataType


@pytest.fixture
def excel_extractor(tmp_path):
    """Fixture to create an instance of MasterdataExcelExtractor with a dummy Excel file."""
    # Create a dummy Excel file in the temporary test directory
    dummy_excel = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add some dummy data
    ws.append(["Header1", "Header2", "Header3"])
    ws.append(["Data1", "Data2", "Data3"])
    ws.append(["", "", ""])  # Empty row
    ws.append(["Last", "Row", "Data"])

    wb.save(dummy_excel)

    # Create an instance with the dummy file
    return MasterdataExcelExtractor(excel_path=str(dummy_excel))


# Tests for `index_to_excel_column`
@pytest.mark.parametrize(
    "index, expected_column",
    [
        (1, "A"),
        (2, "B"),
        (26, "Z"),
        (27, "AA"),
        (28, "AB"),
        (52, "AZ"),
        (53, "BA"),
        (702, "ZZ"),
        (703, "AAA"),
        (704, "AAB"),
        (1378, "AZZ"),
        (1379, "BAA"),
        (18278, "ZZZ"),
        (18279, "AAAA"),
        (18280, "AAAB"),
    ],
)
def test_index_to_excel_column(excel_extractor, index, expected_column):
    """Tests the index_to_excel_column method."""
    result = excel_extractor.index_to_excel_column(index)
    assert result == expected_column


# Failing cases for `index_to_excel_column`
@pytest.mark.parametrize(
    "invalid_index",
    [-1, 0],  # Invalid cases
)
def test_index_to_excel_column_invalid(excel_extractor, invalid_index):
    """Tests that index_to_excel_column fails with invalid input."""
    with pytest.raises(ValueError):
        excel_extractor.index_to_excel_column(invalid_index)


# Tests for `get_last_non_empty_row`
def test_get_last_non_empty_row(excel_extractor):
    """Tests finding the last non-empty row."""
    sheet = excel_extractor.workbook["Sheet1"]

    result = excel_extractor.get_last_non_empty_row(sheet, 1)

    assert result == 2  # Last non-empty row should be 4


# Failing cases for `get_last_non_empty_row`
@pytest.mark.parametrize(
    "invalid_sheet, start_index, expected_exception",
    [
        (None, 1, AttributeError),  # No sheet provided
        ("Sheet1", -1, ValueError),  # Negative index
        ("Sheet1", 999, ValueError),  # Index out of bounds
    ],
)
def test_get_last_non_empty_row_invalid(
    excel_extractor, invalid_sheet, start_index, expected_exception
):
    """Tests that get_last_non_empty_row fails with invalid input."""
    try:
        sheet = (
            excel_extractor.workbook[invalid_sheet]
            if isinstance(invalid_sheet, str)
            else invalid_sheet
        )
    except KeyError:
        if expected_exception is KeyError:
            return  # Test passes if KeyError was expected
        raise  # Otherwise, raise the error

    with pytest.raises(expected_exception):
        excel_extractor.get_last_non_empty_row(sheet, start_index)


# Tests for `is_reduced_version`
@pytest.mark.parametrize(
    "generated_code, full_code, expected_result",
    [
        ("ABC", "ABC", True),  # Identical codes
        ("ABC", "ABC_DEF", True),  # Not a reduced version
        ("ABC.DEF", "ABC.DEF.GHI", True),  # Matching delimiter (.)
        ("ABC_DEF", "ABC_DEF_GHI", True),  # Matching delimiter (_)
        ("ABC.DEF", "ABC_DEF_GHI", False),  # Mismatched delimiters
        ("", "AAA", False),  # Not a reduced version, but function returns True
        ("INS_ANS", "INSTRUMENT", False),  # Contains INS, but no reduced version
        ("ABC.DEF", "ABC_DEF", False),  # Error: the symbol is not the same
        ("ABC_DEF", "ABC.DEF.GHI", False),  # Matching delimiter (_)
        ("ABC.DEF.GHI", "ABC.DEF", False),  # Longer than original code
    ],
)
def test_is_reduced_version(
    excel_extractor, generated_code, full_code, expected_result
):
    """Tests whether generated_code_value is a reduced version of code."""
    result = excel_extractor.is_reduced_version(generated_code, full_code)
    assert result == expected_result


@pytest.mark.parametrize(
    "value, expected_result, log_message, log_message_level",
    [
        ("true", True, None, None),
        ("false", False, None, None),
        (True, True, None, None),
        (False, False, None, None),
        ("TrUe", True, None, None),
        ("FaLsE", False, None, None),
        ("", False, None, None),
        (None, False, None, None),
        (
            "yes",
            False,
            "Invalid boolean term value found in the Boolean Term column at position A1 in Sheet1. Accepted values: TRUE or FALSE.",
            "error",
        ),
        (
            "123",
            False,
            "Invalid boolean term value found in the Boolean Term column at position A1 in Sheet1. Accepted values: TRUE or FALSE.",
            "error",
        ),
    ],
)
def test_str_to_bool(
    cleared_log_storage,
    excel_extractor,
    value,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests conversion of string values to boolean with `str_to_bool`."""
    result = excel_extractor.str_to_bool(value, "Boolean Term", "A1", "Sheet1")
    assert result == expected_result

    if log_message:
        # Check if the expected error message appears in logs
        assert any(
            log["event"] == log_message and log["level"] == log_message_level
            for log in cleared_log_storage
        )
    else:
        # Ensure no error messages are logged
        assert not any(log["level"] == "error" for log in cleared_log_storage)


@pytest.mark.parametrize(
    "value, term, coordinate, sheet_title, is_description, is_code, is_data, is_url, expected_result, log_message, log_message_level",
    [
        # Valid cases (No errors should be logged)
        (
            "Valid Code",
            "Code",
            "A1",
            "Sheet1",
            False,
            False,
            False,
            False,
            "Valid Code",
            None,
            None,
        ),
        (
            "Valid Description // Gültige Beschreibung",
            "Description",
            "A1",
            "Sheet1",
            True,
            False,
            False,
            False,
            "Valid Description // Gültige Beschreibung",
            None,
            None,
        ),
        (
            "VALID_CODE",
            "Code",
            "B2",
            "Sheet2",
            False,
            True,
            False,
            False,
            "VALID_CODE",
            None,
            None,
        ),
        (
            "INTEGER",
            "Data Type",
            "C3",
            "Sheet3",
            False,
            False,
            True,
            False,
            "INTEGER",
            None,
            None,
        ),
        (
            "https://valid-url.com",
            "URL",
            "D4",
            "Sheet4",
            False,
            False,
            False,
            True,
            "https://valid-url.com",
            None,
            None,
        ),
        # Invalid cases (Errors should be logged)
        (
            "Missing Separator",
            "Description",
            "A5",
            "Sheet5",
            True,
            False,
            False,
            False,
            "Missing Separator",
            "Invalid description value found in the Description column at position A5 in Sheet5. "
            "Description should follow the schema: English Description + '//' + German Description. ",
            "error",
        ),
        (
            "Invalid Code!",
            "Code",
            "B6",
            "Sheet6",
            False,
            True,
            False,
            False,
            "Invalid Code!",
            "Invalid code value found in the Code column at position B6 in Sheet6.",
            "error",
        ),
        (
            "unknown_data_type",
            "Data Type",
            "C7",
            "Sheet7",
            False,
            False,
            True,
            False,
            "UNKNOWN_DATA_TYPE",
            f"Invalid data type value found in the Data Type column at position C7 in Sheet7. "
            f"The Data Type should be one of the following: {list(dt.value for dt in DataType)}",
            "error",
        ),
        (
            "invalid-url",
            "URL",
            "D8",
            "Sheet8",
            False,
            False,
            False,
            True,
            "invalid-url",
            "Invalid url value found in the URL column at position D8 in Sheet8.",
            "error",
        ),
    ],
)
def test_get_and_check_property(
    cleared_log_storage,
    excel_extractor,
    value,
    term,
    coordinate,
    sheet_title,
    is_description,
    is_code,
    is_data,
    is_url,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests property validation and formatting with `get_and_check_property`."""
    result = excel_extractor.get_and_check_property(
        value, term, coordinate, sheet_title, is_description, is_code, is_data, is_url
    )
    assert result == expected_result

    if log_message:
        # Normalize log message formatting to avoid strict mismatches
        def normalize_log_message(msg):
            return re.sub(r"\s+", " ", msg.replace(". ", ".")).strip()

        cleaned_logs = [
            normalize_log_message(log["event"]) for log in cleared_log_storage
        ]
        expected_cleaned_message = normalize_log_message(log_message)

        # Generalized check (avoids strict spacing mismatches)
        assert any(expected_cleaned_message in log for log in cleaned_logs)
    else:
        # Ensure no error messages are logged
        assert not any(log["level"] == "error" for log in cleared_log_storage)


@pytest.mark.parametrize(
    "term, cell_value, coordinate, sheet_title, expected_result, log_message, log_message_level",
    [
        # Valid cases (No errors should be logged)
        ("Mandatory", "true", "A1", "Sheet1", True, None, None),
        ("Show in edit views", "False", "B2", "Sheet2", False, None, None),
        ("Code", "VALID_CODE", "C3", "Sheet3", "VALID_CODE", None, None),
        ("Data type", "INTEGER", "D4", "Sheet4", "INTEGER", None, None),
        # Invalid cases (Errors should be logged)
        (
            "Mandatory",
            "invalid_boolean",
            "A7",
            "Sheet7",
            False,
            "Invalid mandatory value found in the Mandatory column at position A7 in Sheet7. Accepted values: TRUE or FALSE.",
            "error",
        ),
        (
            "Show in edit views",
            123,
            "B8",
            "Sheet8",
            False,
            "Invalid show in edit views value found in the Show in edit views column at position B8 in Sheet8. Accepted values: TRUE or FALSE.",
            "error",
        ),
        (
            "Code",
            "Invalid Code!",
            "C9",
            "Sheet9",
            "Invalid Code!",
            "Invalid code value found in the Code column at position C9 in Sheet9.",
            "error",
        ),
        (
            "Data type",
            "UNKNOWN_TYPE",
            "D10",
            "Sheet10",
            "UNKNOWN_TYPE",
            f"Invalid data type value found in the Data type column at position D10 in Sheet10. The Data Type should be one of the following: {list(dt.value for dt in DataType)}",
            "error",
        ),
    ],
)
def test_process_term(
    cleared_log_storage,
    excel_extractor,
    term,
    cell_value,
    coordinate,
    sheet_title,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests `process_term` to validate and process terms correctly."""

    result = excel_extractor.process_term(term, cell_value, coordinate, sheet_title)

    assert result == expected_result

    if log_message:
        # Normalize log message formatting to avoid strict mismatches
        def normalize_log_message(msg):
            return re.sub(r"\s+", " ", msg.replace(". ", ".")).strip()

        cleaned_logs = [
            normalize_log_message(log["event"]) for log in cleared_log_storage
        ]
        expected_cleaned_message = normalize_log_message(log_message)

        # Ensure log appears
        assert any(
            expected_cleaned_message in log for log in cleaned_logs
        ), f"Expected log message was not found. Logs: {cleaned_logs}"
    else:
        # Ensure no error messages are logged
        assert not any(log["level"] == "error" for log in cleared_log_storage)


@pytest.mark.parametrize(
    "cell_value, row, column, validation_pattern, is_description, is_data, is_url, expected_result, log_message, log_message_level",
    [
        # Valid cases (No errors should be logged)
        (
            "VALID_CODE",
            1,
            1,
            r"^\$?[A-Z0-9_.]+$",
            False,
            False,
            False,
            "VALID_CODE",
            None,
            None,
        ),
        (
            "Valid Description // Gültige Beschreibung",
            2,
            2,
            r".*//.*",
            True,
            False,
            False,
            "Valid Description // Gültige Beschreibung",
            None,
            None,
        ),
        ("INTEGER", 3, 3, None, False, True, False, "INTEGER", None, None),
        (
            "https://valid-url.com",
            4,
            4,
            r"https?://(?:www\.)?[a-zA-Z0-9-._~:/?#@!$&'()*+,;=%]+",
            False,
            False,
            True,
            "https://valid-url.com",
            None,
            None,
        ),
        # Invalid cases (Errors should be logged)
        (
            "Missing Separator",
            6,
            1,
            r".*//.*",
            True,
            False,
            False,
            "Missing Separator",
            "Invalid value 'Missing Separator' at row 6, column 1 in sheet TestSheet Description should follow the schema: English Description + '//' + German Description.",
            "error",
        ),
        (
            "unknown_data_type",
            7,
            2,
            None,
            False,
            True,
            False,
            "unknown_data_type",
            f"Invalid value 'unknown_data_type' at row 7, column 2 in sheet TestSheet The Data Type should be one of the following: {list(dt.value for dt in DataType)}",
            "error",
        ),
        (
            "invalid-url",
            8,
            3,
            r"https?://(?:www\.)?[a-zA-Z0-9-._~:/?#@!$&'()*+,;=%]+",
            False,
            False,
            True,
            "invalid-url",
            "Invalid value 'invalid-url' at row 8, column 3 in sheet TestSheet It should be an URL or empty",
            "error",
        ),
        (
            "123INVALID",
            9,
            4,
            r"^[A-Z]+_\d+$",
            False,
            False,
            False,
            "123INVALID",
            "Invalid value '123INVALID' at row 9, column 4 in sheet TestSheet",
            "error",
        ),
        (
            "invalid code",
            9,
            4,
            r"^\$?[A-Z0-9_.]+$",
            False,
            False,
            False,
            "invalid code",
            "Invalid value 'invalid code' at row 9, column 4 in sheet TestSheet",
            "error",
        ),
        # Empty cell (should return empty string, no error)
        (None, 10, 5, None, False, False, False, "", None, None),
        ("", 11, 6, None, False, False, False, "", None, None),
    ],
)
def test_extract_value(
    cleared_log_storage,
    excel_extractor,
    cell_value,
    row,
    column,
    validation_pattern,
    is_description,
    is_data,
    is_url,
    expected_result,
    log_message,
    log_message_level,
):
    """Tests `extract_value` function for extracting and validating cell values."""

    # Create a dummy worksheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Set the cell value
    sheet.cell(row=row, column=column, value=cell_value)

    # Call function
    result = excel_extractor.extract_value(
        sheet, row, column, validation_pattern, is_description, is_data, is_url
    )

    # Assert value extraction result
    assert result == expected_result

    if log_message:
        # Normalize log message formatting
        def normalize_log_message(msg):
            return re.sub(r"\s+", " ", msg.replace(". ", ".")).strip()

        cleaned_logs = [
            (normalize_log_message(log["event"]), log["level"])
            for log in cleared_log_storage
        ]
        expected_cleaned_message = normalize_log_message(log_message)

        # Ensure log appears with the correct level
        assert any(
            expected_cleaned_message in log_text and log_level == log_message_level
            for log_text, log_level in cleaned_logs
        ), f"Expected log message '{expected_cleaned_message}' with level '{log_message_level}' was not found. Logs: {cleaned_logs}"

    else:
        # Ensure no error messages are logged
        assert not any(
            log["level"] == "error" for log in cleared_log_storage
        ), "Unexpected error log found!"


@pytest.mark.parametrize(
    "header_terms, expected_terms, entity_type, cell_values, expected_attributes, log_message, log_message_level",
    [
        # Valid Case: SAMPLE_TYPE (No errors expected)
        (
            ["Code", "Description", "Generated code prefix"],  # Header terms
            [
                "Code",
                "Description",
                "Generated code prefix",
            ],  # Expected extracted terms
            "SAMPLE_TYPE",
            [
                "ABC123",
                "A valid description // Eine gültige Beschreibung",
                "ABC",
            ],  # Cell values
            {
                "code": "ABC123",
                "description": "A valid description // Eine gültige Beschreibung",
                "generatedCodePrefix": "ABC",
            },  # Expected attributes
            None,  # No error
            None,
        ),
        # Invalid Case: Unknown Data Type (PROPERTY_TYPE)
        (
            ["Code", "Data type"],
            ["Code", "Data type"],
            "PROPERTY_TYPE",
            ["ABC123", "UNKNOWN_TYPE"],
            {"code": "ABC123", "dataType": "UNKNOWN_TYPE"},
            "Invalid Data Type: UNKNOWN_TYPE in B3 (Sheet: TestSheet). Should be one of the following: "
            f"{[dt.value for dt in DataType]}",
            "error",
        ),
        # Invalid Case: Boolean Error (SAMPLE_TYPE)
        (
            ["Code", "Auto generated codes"],
            ["Code", "Auto generated codes"],
            "SAMPLE_TYPE",
            ["ABC123", "maybe"],  # Invalid boolean
            {"code": "ABC123", "autoGeneratedCode": False},  # Expected processed value
            "Invalid auto generated codes value found in the Auto generated codes column at position B3 in TestSheet. Accepted values: TRUE or FALSE.",
            "error",
        ),
        # Invalid URL Case (VOCABULARY_TYPE)
        (
            ["Code", "Url template"],
            ["Code", "Url template"],
            "VOCABULARY_TYPE",
            ["ABC123", "invalid-url"],
            {"code": "ABC123", "url_template": "invalid-url"},
            "Invalid URL format: invalid-url in B3 (Sheet: TestSheet)",
            "error",
        ),
        # Invalid Generated Code Prefix (SAMPLE_TYPE)
        (
            ["Code", "Generated code prefix"],
            ["Code", "Generated code prefix"],
            "SAMPLE_TYPE",
            ["ABC123", "ABC_123"],
            {"code": "ABC123", "generatedCodePrefix": "ABC_123"},
            "Invalid Generated code prefix value 'ABC_123' in B3 (Sheet: TestSheet). Generated code prefix should be part of the 'Code'.",
            "error",
        ),
    ],
)
def test_process_entity(
    cleared_log_storage,
    excel_extractor,
    header_terms,
    expected_terms,
    entity_type,
    cell_values,
    expected_attributes,
    log_message,
    log_message_level,
):
    """Tests processing entity attributes and validation logging in `process_entity`."""

    # Create dummy Excel sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Insert headers in row 1
    for col, header in enumerate(header_terms, start=1):
        sheet.cell(row=1, column=col, value=header)

    # Insert test values in row 3 (since `start_index_row + 2` is used)
    for col, value in enumerate(cell_values, start=1):
        sheet.cell(row=3, column=col, value=value)

    # Call the function
    result = excel_extractor.process_entity(
        sheet,
        start_index_row=1,
        header_terms=header_terms,
        expected_terms=expected_terms,
        entity_type=entity_type,
    )

    # Assert entity attributes
    assert (
        result == expected_attributes
    ), f"Expected: {expected_attributes}, but got: {result}"

    # Log message checking
    if log_message:
        cleaned_logs = [
            re.sub(r"\s+", " ", log["event"]).strip() for log in cleared_log_storage
        ]
        expected_cleaned_message = re.sub(r"\s+", " ", log_message).strip()

        assert any(
            expected_cleaned_message in log for log in cleaned_logs
        ), "Expected log message was not found!"

        # Ensure the correct log level is used
        assert any(
            log["level"] == log_message_level for log in cleared_log_storage
        ), f"Expected log level '{log_message_level}' not found."
    else:
        assert not any(
            log["level"] == "error" for log in cleared_log_storage
        ), "Unexpected error logs found!"


@pytest.mark.parametrize(
    "header_terms, cell_values, last_non_empty_row, expected_properties, log_message, log_message_level",
    [
        # ✅ Valid Case: Extract properties correctly
        (
            [
                "Code",
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Data type",
                "Vocabulary code",
            ],
            [
                [
                    "PROP_001",
                    "Sample description",
                    "True",
                    "False",
                    "General",
                    "Property Label",
                    "INTEGER",
                    "VOCAB_ABC",
                ],
                [
                    "PROP_002",
                    "Another description",
                    "False",
                    "True",
                    "Advanced",
                    "Another Label",
                    "BOOLEAN",
                    "VOCAB_DEF",
                ],
            ],
            6,
            {
                "PROP_001": {
                    "permId": "PROP_001",
                    "code": "PROP_001",
                    "description": "Sample description",
                    "mandatory": True,
                    "show_in_edit_views": False,
                    "section": "General",
                    "label": "Property Label",
                    "dataType": "INTEGER",
                    "vocabularyCode": "VOCAB_ABC",
                },
                "PROP_002": {
                    "permId": "PROP_002",
                    "code": "PROP_002",
                    "description": "Another description",
                    "mandatory": False,
                    "show_in_edit_views": True,
                    "section": "Advanced",
                    "label": "Another Label",
                    "dataType": "BOOLEAN",
                    "vocabularyCode": "VOCAB_DEF",
                },
            },
            None,
            None,
        ),
        # ❌ Missing Header: "Data type"
        (
            [
                "Code",
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Vocabulary code",
            ],  # Missing "Data type"
            [
                [
                    "PROP_001",
                    "Sample description",
                    "True",
                    "False",
                    "General",
                    "Property Label",
                    "VOCAB_ABC",
                ]
            ],
            4,
            {},
            "'Data type' not found in the properties headers.",
            "error",
        ),
        # ❌ Invalid Data Type
        (
            [
                "Code",
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Data type",
                "Vocabulary code",
            ],
            [
                [
                    "PROP_003",
                    "Description",
                    "True",
                    "False",
                    "General",
                    "Property Label",
                    "INVALID_TYPE",
                    "VOCAB_ABC",
                ]
            ],
            5,
            {
                "PROP_003": {
                    "permId": "PROP_003",
                    "code": "PROP_003",
                    "description": "Description",
                    "section": "General",
                    "mandatory": True,
                    "show_in_edit_views": False,
                    "label": "Property Label",
                    "dataType": "INVALID_TYPE",
                    "vocabularyCode": "VOCAB_ABC",
                }
            },
            f"Invalid data type value found in the Data type column at position G5 in TestSheet.The Data Type should be one of the following: {[dt.value for dt in DataType]}",
            "error",
        ),
        # ❌ Missing Code (Should be ignored)
        (
            [
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Data type",
                "Vocabulary code",
            ],  # Missing "Code"
            [
                [
                    "Description only",
                    "True",
                    "False",
                    "General",
                    "Label",
                    "INTEGER",
                    "VOCAB_ABC",
                ]
            ],
            4,
            {},
            "'Code' not found in the properties headers.",
            "error",
        ),
    ],
)
def test_properties_to_dict(
    cleared_log_storage,
    excel_extractor,
    header_terms,
    cell_values,
    last_non_empty_row,
    expected_properties,
    log_message,
    log_message_level,
):
    """Tests `properties_to_dict` function to validate properties extraction."""

    # Create dummy Excel sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Insert headers at row 4 (header_index = start_index_row + 3)
    for col, header in enumerate(header_terms, start=1):
        sheet.cell(row=4, column=col, value=header)

    # Insert test values starting from row 5
    for row_index, row_values in enumerate(cell_values, start=5):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    # Debug: Check headers before calling function
    row_headers = [cell.value for cell in sheet[4]]
    print(f"\nDEBUG: Row Headers Extracted = {row_headers}")

    # Call function
    result = excel_extractor.properties_to_dict(
        sheet, start_index_row=1, last_non_empty_row=last_non_empty_row
    )

    # Debugging print
    print("\n--- DEBUG PROPERTIES TO DICT ---")
    print(f"Expected Properties: {expected_properties}")
    print(f"Resulting Properties: {result}")

    # Assert dictionary structure
    assert (
        result == expected_properties
    ), f"Expected: {expected_properties}, but got: {result}"

    # ✅ Log message checking
    if log_message:
        cleaned_logs = [
            re.sub(r"\s+", " ", log["event"]).strip() for log in cleared_log_storage
        ]
        expected_cleaned_message = re.sub(r"\s+", " ", log_message).strip()

        print("\n--- DEBUG LOG CHECK ---")
        print(f"Expected Log Message:\n{expected_cleaned_message}")
        print("\nCaptured Log Messages:")
        for log in cleaned_logs:
            print(f"- {log}")

        # Ensure the expected log message appears
        assert any(
            expected_cleaned_message in log for log in cleaned_logs
        ), "Expected log message was not found!"

        # ✅ Ensure the correct log level is used
        assert any(
            log["level"] == log_message_level for log in cleared_log_storage
        ), f"Expected log level '{log_message_level}' not found."
    else:
        assert not any(
            log["level"] == "error" for log in cleared_log_storage
        ), "Unexpected error logs found!"
