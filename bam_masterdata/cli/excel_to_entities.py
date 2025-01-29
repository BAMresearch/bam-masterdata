import re
from typing import TYPE_CHECKING, Any, Optional, Union

if TYPE_CHECKING:
    from structlog._config import BoundLoggerLazyProxy

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from bam_masterdata.logger import logger


def index_to_excel_column(index: int) -> str:
    """
    Converts a 1-based index to an Excel column name.

    Args:
        index: The 1-based index to convert.

    Returns:
        The corresponding Excel column name.
    """
    column = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        column = chr(65 + remainder) + column
    return column


def get_last_non_empty_row(sheet: Worksheet, start_index: int) -> Optional[int]:
    """
    Finds the last non-empty row before encountering a completely empty row.

    Args:
        sheet: The worksheet object.
        start_index: The row number to start checking from (1-based index).

    Returns:
        The row number of the last non-empty row before an empty row is encountered,
        or None if no non-empty rows are found starting from the given index.
    """
    last_non_empty_row = None
    for row in range(start_index, sheet.max_row + 1):
        if all(
            sheet.cell(row=row, column=col).value in (None, "")
            for col in range(1, sheet.max_column + 1)
        ):
            return last_non_empty_row  # Return the last non-empty row before the current empty row

        last_non_empty_row = row  # Update the last non-empty row

    return last_non_empty_row  # If no empty row is encountered, return the last non-empty row


def is_reduced_version(generated_code_value: str, code: str) -> bool:
    """
    Check if generated_code_value is a reduced version of code.

    Args:
        generated_code_value (str): The potentially reduced code.
        code (str): The original full code.

    Returns:
        bool: True if generated_code_value is a reduced version of code, False otherwise.
    """
    # Check if both are single words (no delimiters)
    if not any(delimiter in code for delimiter in "._") and not any(
        delimiter in generated_code_value for delimiter in "._"
    ):
        return True

    # Determine the delimiter in each string
    code_delimiter = "." if "." in code else "_" if "_" in code else None
    generated_delimiter = (
        "."
        if "." in generated_code_value
        else "_"
        if "_" in generated_code_value
        else None
    )

    # If delimiters don't match, return False
    if code_delimiter != generated_delimiter:
        return False

    # Split both strings using the determined delimiter
    generated_parts = generated_code_value.split(code_delimiter)
    original_parts = code.split(code_delimiter)

    # Ensure both have the same number of parts
    return len(generated_parts) == len(original_parts)


def str_to_bool(
    value: Optional[Union[str, bool]],
    term: str,
    coordinate: str,
    sheet_title: str,
    logger: "BoundLoggerLazyProxy",
) -> bool:
    """
    Converts a string to a boolean value.

    Args:
        value: The string to convert.

    Returns:
        The boolean value.
    """
    # No `value` provided
    if not value:
        return False

    val = str(value).strip().lower()
    if val not in ["true", "false"]:
        logger.error(
            f"Invalid {term.lower()} value found in the {term} column at position {coordinate} in {sheet_title}. Accepted values: TRUE or FALSE."
        )
    return val == "true"


def get_and_check_property(
    value: Optional[Union[str, bool]],
    term: str,
    coordinate: str,
    sheet_title: str,
    logger: "BoundLoggerLazyProxy",
    is_description: bool = False,
    is_code: bool = False,
    is_data: bool = False,
    is_url: bool = False,
) -> str:
    """
    Gets a property and checks its format.

    Args:
        value: The string to convert.

    Returns:
        The property.
    """
    data_types = [
        "INTEGER",
        "REAL",
        "VARCHAR",
        "MULTILINE_VARCHAR",
        "HYPERLINK",
        "BOOLEAN",
        "CONTROLLEDVOCABULARY",
        "XML",
        "TIMESTAMP",
        "DATE",
        "SAMPLE",
        "OBJECT",
    ]

    # No `value` provided
    if not value:
        return ""

    val = str(value)
    error_message = f"Invalid {term.lower()} value found in the {term} column at position {coordinate} in {sheet_title}."
    if is_description:
        if not re.match(r".*//.*", val):
            logger.error(
                error_message
                + "Description should follow the schema: English Description + '//' + German Description. "
            )
    elif is_code:
        if not re.match(r"^\$?[A-Z0-9_.]+$", val):
            logger.error(error_message)
    elif is_data:
        if val not in data_types:
            logger.error(
                error_message
                + "The Data Type should be one of the following: {data_types}"
            )
            val = val.upper()
    elif is_url:
        if not re.match(r"https?://(?:www\.)?[a-zA-Z0-9-._~:/?#@!$&'()*+,;=%]+", val):
            logger.error(error_message)
    else:
        if not re.match(r".*", val):
            logger.error(error_message)
    return val


# Helper function to process each term
def process_term(term, cell_value, coordinate, sheet_title):
    if term in ("Mandatory", "Show in edit views"):
        return str_to_bool(
            value=cell_value,
            term=term,
            coordinate=coordinate,
            sheet_title=sheet_title,
            logger=logger,
        )
    return get_and_check_property(
        value=cell_value,
        term=term,
        coordinate=coordinate,
        sheet_title=sheet_title,
        logger=logger,
        is_code=(term in ["Code", "Vocabulary code"]),
        is_data=(term == "Data type"),
    )


def properties_to_dict(
    sheet: Worksheet, start_index_row: int, last_non_empty_row: int
) -> dict[str, dict[str, Any]]:
    """
    Extracts properties from an Entity type block in the Excel sheet and returns them as a dictionary.

    Args:
        sheet: The worksheet object.
        start_index_row: Row where the current entity type begins (1-based index).
        last_non_empty_row: Row where the current entity type finish (1-based index).

    Returns:
        A dictionary where each key is a property code and the value is a dictionary
        containing the attributes of the property.
    """
    property_dict = {}
    expected_terms = [
        "Code",
        "Description",
        "Mandatory",
        "Show in edit views",
        "Section",
        "Property label",
        "Data type",
        "Vocabulary code",
    ]

    # Determine the header row index
    header_index = start_index_row + 3
    row_headers = [cell.value for cell in sheet[header_index]]

    # Initialize a dictionary to store extracted columns
    extracted_columns: dict[str, list] = {term: [] for term in expected_terms}

    # Extract columns for each expected term
    for term in expected_terms:
        if term not in row_headers:
            log_func = (
                logger.warning
                if term in ("Mandatory", "Show in edit views", "Section")
                else logger.error
            )
            log_func(f"'{term}' not found in the properties headers.")
            continue

        # Get column index and Excel letter
        term_index = row_headers.index(term) + 1
        term_letter = index_to_excel_column(term_index)

        # Extract values from the column
        for cell in sheet[term_letter][header_index:last_non_empty_row]:
            extracted_columns[term].append(
                process_term(term, cell.value, cell.coordinate, sheet.title)
            )

    # Combine extracted values into a dictionary
    for i in range(len(extracted_columns["Code"])):
        property_dict[extracted_columns["Code"][i]] = {
            "permId": extracted_columns["Code"][i],
            "code": extracted_columns["Code"][i],
            "description": extracted_columns["Description"][i],
            "section": extracted_columns["Section"][i],
            "mandatory": extracted_columns["Mandatory"][i],
            "show_in_edit_views": extracted_columns["Show in edit views"][i],
            "label": extracted_columns["Property label"][i],
            "dataType": extracted_columns["Data type"][i],
            "vocabularyCode": extracted_columns["Vocabulary code"][i],
        }

    return property_dict


def terms_to_dict(
    sheet: Worksheet, start_index_row: int, last_non_empty_row: int
) -> dict[str, dict[str, Any]]:
    """
    Extracts terms from a Vocabulary block in the Excel sheet and returns them as a dictionary.

    Args:
        sheet: The worksheet object.
        start_index_row: Row where the current entity type begins (1-based index).
        last_non_empty_row: Row where the current entity type finish (1-based index).

    Returns:
        A dictionary where each key is a vocabulary term code and the value is a dictionary
        containing the attributes of the vocabulary term.
    """
    terms_dict = {}
    expected_terms = ["Code", "Description", "Url template", "Label", "Official"]

    header_index = start_index_row + 3
    row_headers = [cell.value for cell in sheet[header_index]]

    # Initialize a dictionary to store extracted columns
    extracted_columns: dict[str, list] = {term: [] for term in expected_terms}

    # Helper function to process each term
    def process_term(term, cell_value, coordinate, sheet_title):
        if term == "Official":
            return str_to_bool(
                value=cell_value,
                term=term,
                coordinate=coordinate,
                sheet_title=sheet_title,
                logger=logger,
            )
        return get_and_check_property(
            value=cell_value,
            term=term,
            coordinate=coordinate,
            sheet_title=sheet_title,
            logger=logger,
            is_code=(term == "Code"),
            is_url=(term == "Url template"),
        )

    # Extract columns for each expected term
    for term in expected_terms:
        if term not in row_headers:
            logger.warning(f"{term} not found in the properties headers.")
            continue

        # Get column index and Excel letter
        term_index = row_headers.index(term) + 1
        term_letter = index_to_excel_column(term_index)

        # Extract values from the column
        for cell in sheet[term_letter][header_index:last_non_empty_row]:
            extracted_columns[term].append(
                process_term(term, cell.value, cell.coordinate, sheet.title)
            )

    # Combine extracted values into a dictionary
    for i in range(len(extracted_columns["Code"])):
        terms_dict[extracted_columns["Code"][i]] = {
            "permId": extracted_columns["Code"][i],
            "code": extracted_columns["Code"][i],
            "descriptions": extracted_columns["Description"][i],
            "url_template": extracted_columns["Url template"][i],
            "label": extracted_columns["Label"][i],
            "official": extracted_columns["Official"][i],
        }

    return terms_dict


def block_to_entity_dict(
    sheet: Worksheet,
    start_index_row: int,
    last_non_empty_row: int,
    complete_dict: dict[str, Any],
) -> dict[str, Any]:
    """
    Extracts entity attributes from an Excel sheet block and returns them as a dictionary.

    Args:
        sheet: The worksheet object.
        start_index_row: The row where the current entity type begins (1-based index).
        last_non_empty_row: The row where the current entity type finishes (1-based index).
        complete_dict: The dictionary to store the extracted entity attributes.

    Returns:
        A dictionary containing the entity attributes.
    """
    attributes_dict: dict = {}

    # Get the entity type from the specified cell
    entity_type_position = f"A{start_index_row}"
    entity_type = sheet[entity_type_position].value

    # Define the valid entity types
    entity_types = [
        "OBJECT_TYPE",
        "SAMPLE_TYPE",
        "EXPERIMENT_TYPE",
        "DATASET_TYPE",
        "PROPERTY_TYPE",
        "VOCABULARY_TYPE",
    ]

    # Get the header terms from the row below the entity type row
    header_terms = [cell.value for cell in sheet[start_index_row + 1]]

    # Check if the entity type is valid
    if entity_type not in entity_types:
        raise ValueError(
            "The entity type (cell A1) should be one of the following: SAMPLE_TYPE/OBJECT_TYPE, EXPERIMENT_TYPE/COLLECTION_TYPE, DATASET_TYPE, PROPERTY_TYPE, VOCABULARY_TYPE"
        )
    else:
        # Process based on the entity type
        if entity_type == "SAMPLE_TYPE" or entity_type == "OBJECT_TYPE":
            expected_terms = [
                "Code",
                "Description",
                "Validation script",
                "Generated code prefix",
                "Auto generated codes",
            ]

            code_value = ""
            for term in expected_terms:
                if term not in header_terms:
                    logger.error(f"{term} not found in the entity headers.")
                else:
                    # Find the index of the term in the second row
                    term_index = header_terms.index(term)

                    # Check the cell below "Code"
                    if term == "Code":
                        code_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r"^\$?[A-Z0-9_.]+$", str(code_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                            )
                        attributes_dict["permId"] = code_value
                        attributes_dict["code"] = code_value

                    # Check the cell below "Description"
                    elif term == "Description":
                        description_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r".*", str(description_value)):
                            # if not re.match(r".*//.*", str(description_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}. Description should follow the schema: English Description + '//' + German Description. "
                            )
                        attributes_dict["description"] = description_value

                    # Check the cell below "Generated code prefix"
                    elif term == "Generated code prefix":
                        generated_code_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not is_reduced_version(generated_code_value, code_value):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}. The value of 'Generated code prefix' should be a part of the 'Code'."
                            )
                        attributes_dict["generatedCodePrefix"] = generated_code_value

                    # Check the cell below "Validation script"
                    elif term == "Validation script":
                        validation_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if validation_value is not None:
                            if not re.match(
                                r"^[A-Za-z0-9_]+\.py$", str(validation_value)
                            ):
                                logger.error(
                                    f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                                )
                        else:
                            validation_value = ""
                        attributes_dict["validationPlugin"] = validation_value

                    # Check the cell below "Auto generate codes"
                    elif term == "Auto generate codes":
                        cell = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        )
                        auto_generate_value = str_to_bool(
                            value=cell.value,
                            term=term,
                            coordinate=cell.coordinate,
                            sheet_title=sheet.title,
                            logger=logger,
                        )
                        # auto_generate_value = ""
                        # auto_generate_value = auto_generate_value.strip().lower()
                        # if auto_generate_value not in {"true", "false"}:
                        #     logger.error(
                        #         f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                        #     )
                        # auto_generate_value = auto_generate_value == "true"
                        attributes_dict["autoGeneratedCode"] = auto_generate_value

            # Assign the properties dictionary as a field for the entity dictionary
            attributes_dict["properties"] = properties_to_dict(
                sheet, start_index_row, last_non_empty_row
            )

            complete_dict[code_value] = attributes_dict

            # Return the sorted dictionary (by inheritance, using dots "." as criteria for sorting)
            return dict(
                sorted(complete_dict.items(), key=lambda item: item[0].count("."))
            )

        elif entity_type == "EXPERIMENT_TYPE" or entity_type == "DATASET_TYPE":
            expected_terms = ["Code", "Description", "Validation script"]
            for term in expected_terms:
                if term not in header_terms:
                    logger.error(f"{term} not found in the second row.")
                else:
                    # Find the index of the term in the second row
                    term_index = header_terms.index(term)

                    # Check the cell below "Code"
                    if term == "Code":
                        code_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r"^\$?[A-Z0-9_.]+$", str(code_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                            )
                        attributes_dict["permId"] = code_value
                        attributes_dict["code"] = code_value

                    # Check the cell below "Description"
                    elif term == "Description":
                        description_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r".*", str(description_value)):
                            # if not re.match(r".*//.*", str(description_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}. Description should follow the schema: English Description + '//' + German Description. "
                            )
                        attributes_dict["description"] = description_value

                    # Check the cell below "Validation script"
                    elif term == "Validation script":
                        validation_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if validation_value is not None:
                            if not re.match(
                                r"^[A-Za-z0-9_]+\.py$", str(validation_value)
                            ):
                                logger.error(
                                    f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                                )
                        else:
                            validation_value = ""
                        attributes_dict["validationPlugin"] = validation_value

            # Assign the properties dictionary as a field for the entity dictionary
            attributes_dict["properties"] = properties_to_dict(
                sheet, start_index_row, last_non_empty_row
            )

            complete_dict[code_value] = attributes_dict

            # Return the sorted dictionary (by inheritance, using dots "." as criteria for sorting)
            return dict(
                sorted(complete_dict.items(), key=lambda item: item[0].count("."))
            )

        elif entity_type == "PROPERTY_TYPE":
            expected_terms = [
                "Code",
                "Description",
                "Property label",
                "Data type",
                "Vocabulary code",
                "Metadata",
                "Dynamic script",
            ]
            for term in expected_terms:
                if term not in header_terms:
                    logger.error(f"{term} not found in the second row.")
                else:
                    # Find the index of the term in the second row
                    term_index = header_terms.index(term)

                    # Check the cell below "Code"
                    if term == "Code":
                        code_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r"^\$?[A-Z0-9_.]+$", str(code_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                            )
                        attributes_dict["permId"] = code_value
                        attributes_dict["code"] = code_value

                    # Check the cell below "Description"
                    elif term == "Description":
                        description_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r".*", str(description_value)):
                            # if not re.match(r".*//.*", str(description_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}. Description should follow the schema: English Description + '//' + German Description. "
                            )
                        attributes_dict["description"] = description_value

                    # Check the cell below "Property label"
                    elif term == "Property label":
                        property_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r".*", str(property_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                            )
                        attributes_dict["label"] = property_value

                    # Check the cell below "Data type"
                    elif term == "Data type":
                        data_types = [
                            "INTEGER",
                            "REAL",
                            "VARCHAR",
                            "MULTILINE_VARCHAR",
                            "HYPERLINK",
                            "BOOLEAN",
                            "CONTROLLEDVOCABULARY",
                            "XML",
                            "TIMESTAMP",
                            "DATE",
                            "SAMPLE",
                            "OBJECT",
                        ]
                        data_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if str(data_value) not in data_types:
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}. The Data Type should be one of the following: {data_types}"
                            )
                        attributes_dict["dataType"] = data_value

                    # Check the cell below "Vocabulary code"
                    elif term == "Vocabulary code":
                        vocabulary_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if vocabulary_value is not None:
                            if not re.match(r"^\$?[A-Z0-9_.]+$", str(vocabulary_value)):
                                logger.error(
                                    f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                                )
                        else:
                            vocabulary_value = ""
                        attributes_dict["vocabularyCode"] = vocabulary_value

                    # Check the cell below "Data type"
                    elif term == "Metadata":
                        metadata_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if metadata_value is not None:
                            if not re.match(r".*", str(metadata_value)):
                                logger.error(
                                    f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                                )
                        else:
                            metadata_value = ""
                        attributes_dict["metadata"] = metadata_value

                    # Check the cell below "Dynamic script"
                    elif term == "Dynamic script":
                        plugin_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if plugin_value is not None:
                            if not re.match(r"^[A-Za-z0-9_]+\.py$", str(plugin_value)):
                                logger.error(
                                    f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                                )
                        else:
                            plugin_value = ""
                        attributes_dict["plugin"] = plugin_value

            complete_dict[code_value] = attributes_dict

            # Return the sorted dictionary (by inheritance, using dots "." as criteria for sorting)
            return dict(
                sorted(complete_dict.items(), key=lambda item: item[0].count("."))
            )

        elif entity_type == "VOCABULARY_TYPE":
            expected_terms = ["Code", "Description", "Url template"]
            for term in expected_terms:
                if term not in header_terms:
                    logger.error(f"{term} not found in the second row.")
                else:
                    # Find the index of the term in the second row
                    term_index = header_terms.index(term)

                    # Check the cell below "Code"
                    if term == "Code":
                        code_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r"^\$?[A-Z0-9_.]+$", str(code_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}"
                            )
                        attributes_dict["permId"] = code_value
                        attributes_dict["code"] = code_value

                    # Check the cell below "Description"
                    elif term == "Description":
                        description_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if not re.match(r".*", str(description_value)):
                            # if not re.match(r".*//.*", str(description_value)):
                            logger.error(
                                f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}. Description should follow the schema: English Description + '//' + German Description. "
                            )
                        attributes_dict["description"] = description_value

                    # Check the cell below "URL Template"
                    elif term == "Url template":
                        url_value = sheet.cell(
                            row=start_index_row + 2, column=term_index + 1
                        ).value
                        if url_value is not None:
                            if not re.match(
                                r"https?://(?:www\.)?[a-zA-Z0-9-._~:/?#@!$&'()*+,;=%]+",
                                str(url_value),
                            ):
                                logger.error(
                                    f"Invalid {term.lower()} value found in the {term} value for entity {code_value} at row {start_index_row + 2}. It should be an URL or empty"
                                )
                        else:
                            url_value = ""
                        attributes_dict["url_template"] = url_value

            # Assign the terms dictionary as a field for the vocabulary dictionary
            attributes_dict["terms"] = terms_to_dict(
                sheet, start_index_row, last_non_empty_row
            )

            complete_dict[code_value] = attributes_dict

            # Return the sorted dictionary (by inheritance, using dots "." as criteria for sorting)
            return dict(
                sorted(complete_dict.items(), key=lambda item: item[0].count("."))
            )
        return attributes_dict


def excel_to_entities(
    excel_path: str, output_directory: str = "./artifacts/tmp/"
) -> dict[str, dict[str, Any]]:
    """
    Extracts entities from an Excel file and returns them as a dictionary.

    Args:
        excel_path: The path to the Excel file.
        output_directory: The directory to store the output files.

    Returns:
        A dictionary where each key is a normalized sheet name and the value is a dictionary
        containing the extracted entities.
    """
    sheets_dict: dict[str, dict[str, Any]] = {}

    # Load the workbook and get the sheet names
    workbook = openpyxl.load_workbook(excel_path)
    sheet_names = workbook.sheetnames

    for i, sheet_name in enumerate(sheet_names):
        normalized_sheet_name = sheet_name.lower().replace(" ", "_")

        sheet = workbook[sheet_name]
        start_row = 1

        sheets_dict[normalized_sheet_name] = {}

        while start_row <= sheet.max_row:
            # Find the last non-empty row of the current block
            last_non_empty_row = get_last_non_empty_row(sheet, start_row)

            # Check if we've reached the end of the sheet or found two consecutive empty rows
            if last_non_empty_row is None:
                if i == len(sheet_names) - 1:  # Check if it's the last sheet
                    logger.info(
                        f"Last sheet {sheet_name} processed. End of the file reached."
                    )
                else:
                    logger.info(
                        f"End of the current sheet {sheet_name} reached. Switching to next sheet..."
                    )
                break

            # Process the block (from start_row to last_non_empty_row)
            sheets_dict[normalized_sheet_name] = block_to_entity_dict(
                sheet, start_row, last_non_empty_row, sheets_dict[normalized_sheet_name]
            )

            # Update start_row to the row after the empty row
            start_row = last_non_empty_row + 1
            while start_row <= sheet.max_row and all(
                sheet.cell(row=start_row, column=col).value in (None, "")
                for col in range(1, sheet.max_column + 1)
            ):
                start_row += 1

            # Check if there are two consecutive empty rows
            if start_row > sheet.max_row or all(
                sheet.cell(row=start_row, column=col).value in (None, "")
                for col in range(1, sheet.max_column + 1)
            ):
                if i == len(sheet_names) - 1:  # Check if it's the last sheet
                    logger.info(
                        f"Last sheet {sheet_name} processed. End of the file reached."
                    )
                else:
                    logger.info(
                        f"End of the current sheet {sheet_name} reached. Switching to next sheet..."
                    )
                break

    return sheets_dict
