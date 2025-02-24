import json
import os
import re
from typing import TYPE_CHECKING, Any, Optional, Union

from bam_masterdata.utils import load_validation_rules

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

import openpyxl

from bam_masterdata.logger import logger
from bam_masterdata.metadata.definitions import DataType


class MasterdataExcelExtractor:
    # TODO move these validation rules to a separate json
    VALIDATION_RULES: dict[str, dict[str, dict[str, Any]]] = {}

    def __init__(self, excel_path: str, **kwargs):
        """Initialize the MasterdataExtractor."""
        self.excel_path = excel_path
        self.row_cell_info = kwargs.get("row_cell_info", False)
        self.workbook = openpyxl.load_workbook(excel_path)
        self.logger = kwargs.get("logger", logger)

        # Load validation rules at initialization
        if not MasterdataExcelExtractor.VALIDATION_RULES:
            self.VALIDATION_RULES = load_validation_rules(self.logger)

    def index_to_excel_column(self, index: int) -> str:
        """
        Converts a 1-based index to an Excel column name.

        Args:
            index: The 1-based index to convert.

        Returns:
            The corresponding Excel column name.
        """
        if not index >= 1:
            raise ValueError("Index must be a positive integer starting from 1.")

        column = ""
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            column = chr(65 + remainder) + column
        return column

    def get_last_non_empty_row(
        self, sheet: "Worksheet", start_index: int
    ) -> Optional[int]:
        """
        Finds the last non-empty row before encountering a completely empty row.

        Args:
            sheet: The worksheet object.
            start_index: The row number to start checking from (1-based index).

        Returns:
            The row number of the last non-empty row before an empty row is encountered,
            or None if no non-empty rows are found starting from the given index.
        """
        if start_index < 1 or start_index > sheet.max_row:
            raise ValueError(
                f"Invalid start index: {start_index}. It must be between 1 and {sheet.max_row}."
            )

        last_non_empty_row = None
        for row in range(start_index, sheet.max_row + 1):
            if all(
                sheet.cell(row=row, column=col).value in (None, "")
                for col in range(1, sheet.max_column + 1)
            ):
                return last_non_empty_row  # Return the last non-empty row before the current empty row

            last_non_empty_row = row  # Update the last non-empty row

        return last_non_empty_row  # If no empty row is encountered, return the last non-empty row

    def is_reduced_version(self, generated_code_value: str, code: str) -> bool:
        """
        Check if generated_code_value is a reduced version of code.

        Args:
            generated_code_value (str): The potentially reduced code.
            code (str): The original full code.

        Returns:
            bool: True if generated_code_value is a reduced version of code, False otherwise.
        """
        if generated_code_value == "" or code == "":
            return False

        if code.startswith(generated_code_value):
            return True

        # Check if both are single words (no delimiters)
        if not any(delimiter in code for delimiter in "._") and not any(
            delimiter in generated_code_value for delimiter in "._"
        ):
            return True

        # Determine the delimiter in each string
        code_delimiter = "." if "." in code else "_" if "_" in code else None
        generated_delimiter = (
            "."
            if "." in generated_code_value
            else "_"
            if "_" in generated_code_value
            else None
        )

        # If delimiters don't match, return False
        if code_delimiter != generated_delimiter:
            return False

        # Split both strings using the determined delimiter
        generated_parts = generated_code_value.split(code_delimiter)
        original_parts = code.split(code_delimiter)

        # Ensure both have the same number of parts
        return len(generated_parts) == len(original_parts)

    def str_to_bool(
        self,
        value: Optional[Union[str, bool]],
        term: str,
        coordinate: str,
        sheet_title: str,
    ) -> bool:
        """
        Converts a string to a boolean value.

        Args:
            value: The string to convert.

        Returns:
            The boolean value.
        """
        # No `value` provided
        if not value:
            return False

        val = str(value).strip().lower()
        if val not in ["true", "false"]:
            self.logger.error(
                f"Invalid {term.lower()} value found in the {term} column at position {coordinate} in {sheet_title}. Accepted values: TRUE or FALSE.",
                term=term,
                cell_value=val,
                cell_coordinate=coordinate,
                sheet_title=sheet_title,
            )
        return val == "true"

    def get_and_check_property(
        self,
        value: Optional[Union[str, bool]],
        term: str,
        coordinate: str,
        sheet_title: str,
        is_description: bool = False,
        is_code: bool = False,
        is_data: bool = False,
        is_url: bool = False,
    ) -> str:
        """
        Gets a property and checks its format.

        Args:
            value: The string to convert.

        Returns:
            The property.
        """

        # No `value` provided
        if not value:
            return ""

        val = str(value)
        error_message = f"Invalid {term.lower()} value found in the {term} column at position {coordinate} in {sheet_title}."
        if is_description:
            if not re.match(r".*//.*", val):
                self.logger.error(
                    error_message
                    + "Description should follow the schema: English Description + '//' + German Description. ",
                    term=term,
                    cell_value=val,
                    cell_coordinate=coordinate,
                    sheet_title=sheet_title,
                )
        elif is_code:
            if not re.match(r"^\$?[A-Z0-9_.]+$", val):
                self.logger.error(
                    error_message,
                    term=term,
                    cell_value=val,
                    cell_coordinate=coordinate,
                    sheet_title=sheet_title,
                )
        elif is_data:
            if val not in [dt.value for dt in DataType]:
                self.logger.error(
                    error_message
                    + f"The Data Type should be one of the following: {[dt.value for dt in DataType]}",
                    term=term,
                    cell_value=val,
                    cell_coordinate=coordinate,
                    sheet_title=sheet_title,
                )
                val = val.upper()
        elif is_url:
            if not re.match(
                r"https?://(?:www\.)?[a-zA-Z0-9-._~:/?#@!$&'()*+,;=%]+", val
            ):
                self.logger.error(
                    error_message,
                    term=term,
                    cell_value=val,
                    cell_coordinate=coordinate,
                    sheet_title=sheet_title,
                )
        else:
            if not re.match(r".*", val):
                self.logger.error(
                    error_message,
                    term=term,
                    cell_value=val,
                    cell_coordinate=coordinate,
                    sheet_title=sheet_title,
                )
        return val

    # Helper function to process each term
    def process_term(
        self, term: str, cell_value: Any, coordinate: str, sheet_title: str
    ) -> Any:
        """
        Processes a term by converting it to a boolean if necessary or checking its validity.

        Args:
            term: The term being processed.
            cell_value: The value of the cell.
            coordinate: The coordinate of the cell in the sheet.
            sheet_title: The title of the sheet.

        Returns:
            The processed value, either as a boolean or the original value after validation.
        """
        # Check if the term is a boolean type
        if term in ("Mandatory", "Show in edit views"):
            return self.str_to_bool(
                value=cell_value,
                term=term,
                coordinate=coordinate,
                sheet_title=sheet_title,
            )
        # Check and validate the property
        return self.get_and_check_property(
            value=cell_value,
            term=term,
            coordinate=coordinate,
            sheet_title=sheet_title,
            is_code=(term in ["Code", "Vocabulary code"]),
            is_data=(term == "Data type"),
        )

    def extract_value(
        self,
        sheet: "Worksheet",
        row: int,
        column: int,
        validation_pattern: str = None,
        is_description: bool = False,
        is_data: bool = False,
        is_url: bool = False,
    ) -> str:
        """
        Extracts and validates a value from a specified cell in the Excel sheet.

        Args:
            sheet: The worksheet object.
            row: The row number of the cell (1-based index).
            column: The column number of the cell (1-based index).
            validation_pattern: Optional regex pattern to validate the cell value.
            is_description: Flag indicating if the value is a description.
            is_data: Flag indicating if the value is a data type.
            is_url: Flag indicating if the value is a URL.

        Returns:
            The extracted and validated cell value as a string. Returns an empty string if the value is invalid or not provided.
        """
        value = sheet.cell(row=row, column=column).value

        # No `value` provided
        if not value:
            return ""

        validated = (
            bool(re.match(validation_pattern, str(value)))
            if validation_pattern
            else True
        )
        error_message = f"Invalid value '{value}' at row {row}, column {column} in sheet {sheet.title}"

        if is_description:
            error_message += " Description should follow the schema: English Description + '//' + German Description."
        elif is_data:
            validated = str(value) in [dt.value for dt in DataType]
            error_message += f" The Data Type should be one of the following: {[dt.value for dt in DataType]}"
        elif is_url:
            error_message += " It should be an URL or empty"

        if not validated:
            self.logger.error(
                error_message,
                cell_value=value,
                sheet_title=sheet.title,
                row=row,
                column=column,
            )

        return value or ""

    def process_entity(
        self,
        sheet: "Worksheet",
        start_index_row: int,
        header_terms: list[str],
        expected_terms: list[str],
        entity_type: str,
    ) -> dict[str, Any]:
        """
        Process an entity type block in the Excel sheet and return its attributes as a dictionary.

        Args:
            sheet: The worksheet object.
            start_index_row: The row where the current entity type begins (1-based index).
            header_terms: List of header terms in the entity block.
            expected_terms: List of expected terms to extract from the entity block.
            entity_type: The type of the entity (e.g., SAMPLE_TYPE, OBJECT_TYPE).

        Returns:
            A dictionary containing the attributes of the entity.
        """
        attributes: dict = {}
        cell_value: Any = ""

        for term in expected_terms:
            if term not in header_terms:
                self.logger.error(f"{term} not found in the headers.", term=term)
            else:
                term_index = header_terms.index(term)
                cell = sheet.cell(row=start_index_row + 2, column=term_index + 1)
                cell_value = self.extract_value(
                    sheet,
                    start_index_row + 2,
                    term_index + 1,
                    self.VALIDATION_RULES[entity_type][term].get("pattern"),
                )

                # Handle boolean conversion
                if self.VALIDATION_RULES[entity_type][term].get("is_bool"):
                    cell_value = self.str_to_bool(
                        value=cell_value,
                        term=term,
                        coordinate=cell.coordinate,
                        sheet_title=sheet.title,
                    )

                # Handle data type validation
                elif self.VALIDATION_RULES[entity_type][term].get("is_data"):
                    if cell_value not in [dt.value for dt in DataType]:
                        self.logger.error(
                            f"Invalid Data Type: {cell_value} in {cell.coordinate} (Sheet: {sheet.title}). Should be one of the following: {[dt.value for dt in DataType]}",
                            term=term,
                            cell_value=cell_value,
                            cell_coordinate=cell.coordinate,
                            sheet_title=sheet.title,
                        )

                # Handle additional validation for "Generated code prefix"
                elif (
                    self.VALIDATION_RULES[entity_type][term].get("extra_validation")
                    == "is_reduced_version"
                ):
                    if not self.is_reduced_version(
                        cell_value, attributes.get("code", "")
                    ):
                        self.logger.warning(
                            f"Invalid {term} value '{cell_value}' in {cell.coordinate} (Sheet: {sheet.title}). "
                            f"Generated code prefix should be part of the 'Code' {attributes.get('code', '')}.",
                            term=term,
                            cell_value=cell_value,
                            cell_coordinate=cell.coordinate,
                            sheet_title=sheet.title,
                        )

                # Handle validation script (allows empty but must match pattern if provided)
                elif (
                    self.VALIDATION_RULES[entity_type][term].get("allow_empty")
                    and not cell_value
                ):
                    cell_value = ""

                # Handle URL template validation (allows empty but must be a valid URL)
                elif (
                    self.VALIDATION_RULES[entity_type][term].get("is_url")
                    and cell_value
                ):
                    url_pattern = self.VALIDATION_RULES[entity_type][term].get(
                        "pattern"
                    )
                    if not re.match(url_pattern, str(cell_value)):
                        self.logger.error(
                            f"Invalid URL format: {cell_value} in {cell.coordinate} (Sheet: {sheet.title})",
                            cell_value=cell_value,
                            cell_coordinate=cell.coordinate,
                            sheet_title=sheet.title,
                        )

                # Add the extracted value to the attributes dictionary
                attributes[self.VALIDATION_RULES[entity_type][term].get("key")] = (
                    cell_value
                )

        if self.row_cell_info:
            attributes["row_location"] = f"A{start_index_row}"
        return attributes

    def properties_to_dict(
        self, sheet: "Worksheet", start_index_row: int, last_non_empty_row: int
    ) -> dict[str, dict[str, Any]]:
        """
        Extracts properties from an Entity type block in the Excel sheet and returns them as a dictionary.

        Args:
            sheet: The worksheet object.
            start_index_row: Row where the current entity type begins (1-based index).
            last_non_empty_row: Row where the current entity type finish (1-based index).

        Returns:
            A dictionary where each key is a property code and the value is a dictionary
            containing the attributes of the property.
        """
        property_dict: dict = {}
        expected_terms = [
            "Code",
            "Description",
            "Mandatory",
            "Show in edit views",
            "Section",
            "Property label",
            "Data type",
            "Vocabulary code",
        ]

        # Determine the header row index
        header_index = start_index_row + 3
        row_headers = [cell.value for cell in sheet[header_index]]

        # Initialize a dictionary to store extracted columns
        extracted_columns: dict[str, list] = {term: [] for term in expected_terms}
        if self.row_cell_info:
            extracted_columns["Row location"] = []

        # Extract columns for each expected term
        for term in expected_terms:
            if term not in row_headers:
                log_func = (
                    self.logger.warning
                    if term in ("Mandatory", "Show in edit views", "Section")
                    else self.logger.error
                )
                log_func(f"'{term}' not found in the properties headers.", term=term)
                continue

            # Get column index and Excel letter
            term_index = row_headers.index(term) + 1
            term_letter = self.index_to_excel_column(term_index)

            # Extract values from the column
            for i in range(header_index + 1, last_non_empty_row + 1):
                cell = sheet[term_letter + str(i)]
                extracted_columns[term].append(
                    self.process_term(term, cell.value, cell.coordinate, sheet.title)
                )
                if self.row_cell_info:
                    extracted_columns["Row location"].append(term_letter + str(i))
            # for cell in sheet[term_letter][header_index:last_non_empty_row]:
            #     extracted_columns[term].append(
            #         self.process_term(term, cell.value, cell.coordinate, sheet.title)
            #     )

        # Combine extracted values into a dictionary
        for i in range(len(extracted_columns["Code"])):
            property_dict[extracted_columns["Code"][i]] = {
                "permId": extracted_columns["Code"][i],
                "code": extracted_columns["Code"][i],
                "description": extracted_columns["Description"][i],
                "section": extracted_columns["Section"][i],
                "mandatory": extracted_columns["Mandatory"][i],
                "show_in_edit_views": extracted_columns["Show in edit views"][i],
                "label": extracted_columns["Property label"][i],
                "dataType": extracted_columns["Data type"][i],
                "vocabularyCode": extracted_columns["Vocabulary code"][i],
            }
            if self.row_cell_info:
                property_dict[extracted_columns["Code"][i]]["row_location"] = (
                    extracted_columns["Row location"][i]
                )

        return property_dict

    def terms_to_dict(
        self, sheet: "Worksheet", start_index_row: int, last_non_empty_row: int
    ) -> dict[str, dict[str, Any]]:
        """
        Extracts terms from a Vocabulary block in the Excel sheet and returns them as a dictionary.

        Args:
            sheet: The worksheet object.
            start_index_row: Row where the current entity type begins (1-based index).
            last_non_empty_row: Row where the current entity type finish (1-based index).

        Returns:
            A dictionary where each key is a vocabulary term code and the value is a dictionary
            containing the attributes of the vocabulary term.
        """
        terms_dict = {}
        expected_terms = ["Code", "Description", "Url template", "Label", "Official"]

        header_index = start_index_row + 3
        row_headers = [cell.value for cell in sheet[header_index]]

        # Initialize a dictionary to store extracted columns
        extracted_columns: dict[str, list] = {term: [] for term in expected_terms}

        # Helper function to process each term
        def process_term(term, cell_value, coordinate, sheet_title):
            if term == "Official":
                return self.str_to_bool(
                    value=cell_value,
                    term=term,
                    coordinate=coordinate,
                    sheet_title=sheet_title,
                )
            return self.get_and_check_property(
                value=cell_value,
                term=term,
                coordinate=coordinate,
                sheet_title=sheet_title,
                is_code=(term == "Code"),
                is_url=(term == "Url template"),
            )

        # Extract columns for each expected term
        for term in expected_terms:
            if term not in row_headers:
                self.logger.warning(
                    f"{term} not found in the properties headers.", term=term
                )
                continue

            # Get column index and Excel letter
            term_index = row_headers.index(term) + 1
            term_letter = self.index_to_excel_column(term_index)

            # Extract values from the column
            for cell in sheet[term_letter][header_index:last_non_empty_row]:
                extracted_columns[term].append(
                    process_term(term, cell.value, cell.coordinate, sheet.title)
                )

        # Combine extracted values into a dictionary
        for i in range(len(extracted_columns["Code"])):
            terms_dict[extracted_columns["Code"][i]] = {
                "permId": extracted_columns["Code"][i],
                "code": extracted_columns["Code"][i],
                "descriptions": extracted_columns["Description"][i],
                "url_template": extracted_columns["Url template"][i],
                "label": extracted_columns["Label"][i],
                "official": extracted_columns["Official"][i],
            }

        return terms_dict

    def block_to_entity_dict(
        self,
        sheet: "Worksheet",
        start_index_row: int,
        last_non_empty_row: int,
        complete_dict: dict[str, Any],
    ) -> dict[str, Any]:
        """
        Extracts entity attributes from an Excel sheet block and returns them as a dictionary.
        """
        attributes_dict: dict = {}

        # Get the entity type
        entity_type = sheet[f"A{start_index_row}"].value
        if entity_type not in self.VALIDATION_RULES:
            raise ValueError(f"Invalid entity type: {entity_type}")

        # Get the header terms
        header_terms = [cell.value for cell in sheet[start_index_row + 1]]

        # Process entity data using the helper function
        attributes_dict = self.process_entity(
            sheet,
            start_index_row,
            header_terms,
            list(self.VALIDATION_RULES[entity_type].keys()),
            entity_type,
        )

        # Extract additional attributes if necessary
        if entity_type in {
            "SAMPLE_TYPE",
            "OBJECT_TYPE",
            "EXPERIMENT_TYPE",
            "DATASET_TYPE",
        }:
            attributes_dict["properties"] = (
                self.properties_to_dict(sheet, start_index_row, last_non_empty_row)
                or {}
            )

        elif entity_type == "VOCABULARY_TYPE":
            attributes_dict["terms"] = (
                self.terms_to_dict(sheet, start_index_row, last_non_empty_row) or {}
            )

        # Add the entity to the complete dictionary
        complete_dict[attributes_dict["code"]] = attributes_dict

        # Return sorted dictionary
        return dict(sorted(complete_dict.items(), key=lambda item: item[0].count(".")))

    def excel_to_entities(self) -> dict[str, dict[str, Any]]:
        """
        Extracts entities from an Excel file and returns them as a dictionary.

        Args:
            output_directory: The directory to store the output files.

        Returns:
            A dictionary where each key is a normalized sheet name and the value is a dictionary
            containing the extracted entities. Returns an empty dictionary if all sheets are empty.
        """
        sheets_dict: dict[str, dict[str, Any]] = {}
        sheet_names = self.workbook.sheetnames
        has_content = False  # Track if any sheet has valid content

        for i, sheet_name in enumerate(sheet_names):
            normalized_sheet_name = sheet_name.lower().replace(" ", "_")
            sheet = self.workbook[sheet_name]
            start_row = 1

            # **Check if the sheet is empty**
            if all(
                sheet.cell(row=row, column=col).value in (None, "")
                for row in range(1, sheet.max_row + 1)
                for col in range(1, sheet.max_column + 1)
            ):
                self.logger.info(f"Skipping empty sheet: {sheet_name}")
                continue  # Move to the next sheet

            sheets_dict[normalized_sheet_name] = {}

            while start_row <= sheet.max_row:
                # Find the last non-empty row of the current block
                last_non_empty_row = self.get_last_non_empty_row(sheet, start_row)

                # Check if we've reached the end of the sheet or found two consecutive empty rows
                if last_non_empty_row is None:
                    if i == len(sheet_names) - 1:  # Check if it's the last sheet
                        self.logger.info(
                            f"Last sheet {sheet_name} processed. End of the file reached."
                        )
                    else:
                        self.logger.info(
                            f"End of the current sheet {sheet_name} reached. Switching to next sheet..."
                        )
                    break

                # **Process the block and mark that we found content**
                sheets_dict[normalized_sheet_name] = self.block_to_entity_dict(
                    sheet,
                    start_row,
                    last_non_empty_row,
                    sheets_dict[normalized_sheet_name],
                )
                has_content = True  # Found at least one valid entity block

                # Update start_row to the row after the empty row
                consecutive_empty_rows = 0  # Track empty row count
                start_row = last_non_empty_row + 1

                while start_row <= sheet.max_row:
                    is_row_empty = all(
                        sheet.cell(row=start_row, column=col).value in (None, "")
                        for col in range(1, sheet.max_column + 1)
                    )

                    if is_row_empty:
                        consecutive_empty_rows += 1
                        start_row += 1  # Move to the next row

                        if consecutive_empty_rows >= 2:  # Stop processing the sheet
                            return sheets_dict  # Ensure function exits immediately!
                    else:
                        consecutive_empty_rows = (
                            0  # Reset counter if a non-empty row is found
                        )

                        last_non_empty_row = self.get_last_non_empty_row(
                            sheet, start_row
                        )
                        sheets_dict[normalized_sheet_name] = self.block_to_entity_dict(
                            sheet,
                            start_row,
                            last_non_empty_row,
                            sheets_dict[normalized_sheet_name],
                        )

                        start_row = (
                            last_non_empty_row + 1
                        )  # Move to the next entity block

                # Check if there are two consecutive empty rows
                if start_row > sheet.max_row or all(
                    sheet.cell(row=start_row, column=col).value in (None, "")
                    for col in range(1, sheet.max_column + 1)
                ):
                    if i == len(sheet_names) - 1:  # Check if it's the last sheet
                        self.logger.info(
                            f"Last sheet {sheet_name} processed. End of the file reached."
                        )
                    else:
                        self.logger.info(
                            f"End of the current sheet {sheet_name} reached. Switching to next sheet..."
                        )
                    break

        # **If no sheets had content, return an empty dictionary**
        if not has_content:
            self.logger.warning(
                "No valid data found in any sheets. Returning empty dictionary."
            )
            return {}

        return sheets_dict
